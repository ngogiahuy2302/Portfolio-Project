import os
import requests
from bs4 import BeautifulSoup
import pandas as pd
from openpyxl import load_workbook, Workbook

def load_skus_from_excel(excel_path):
    """ Load SKUs from an Excel file into a list """
    df = pd.read_excel(excel_path)
    return df['Variant SKU'].tolist()

def check_sku_existence(sku_list, search_url, sub_folder1):
    """ Check each SKU on a website and download the page if found """
    found = []
    not_found = []
    for sku in sku_list:
        response = requests.get(search_url, params={'text': sku})
        if response.status_code == 200:
            soup = BeautifulSoup(response.text, 'html.parser')
            result_div = soup.find('div', class_='item homologation-productResult-0')
            if result_div:
                link = result_div.find('a', href=True)
                if link:
                    product_url = 'https://www.tefal.com.sg' + link['href'] if not link['href'].startswith('http') else link['href']
                    product_response = requests.get(product_url)
                    if product_response.status_code == 200:
                        found.append(sku)
                        sku_folder = os.path.join(sub_folder1, sku)
                        os.makedirs(sku_folder, exist_ok=True)
                        with open(os.path.join(sku_folder, 'SourcecodeHTML.txt'), 'w', encoding='utf-8') as html_file:
                            html_file.write(product_response.text)
                else:
                    print(f"Link not found for SKU: {sku}")
                    not_found.append(sku)
            else:
                not_found.append(sku)
    return found, not_found

def extract_and_clean_data(sub_folder1):
    """ Extract and clean text from downloaded HTML content """
    for subdir, dirs, files in os.walk(sub_folder1):
        for file in files:
            if file.endswith("SourcecodeHTML.txt"):
                filepath = os.path.join(subdir, file)
                with open(filepath, 'r', encoding='utf-8') as html_file:
                    soup = BeautifulSoup(html_file.read(), 'html.parser')
                essential_list = soup.find('div', {'id': 'EssentialList'})
                if essential_list:
                    items = essential_list.find_all('li')
                    cleaned_text = ' '.join(f"{item.h3.text} {item.p.text}" for item in items if item.h3 and item.p)
                    with open(os.path.join(subdir, 'Textprepared.txt'), 'w', encoding='utf-8') as text_file:
                        text_file.write(cleaned_text)

def update_excel_with_description(output_excel_path, sub_folder1):
    """ Update or create an Excel file with extracted descriptions, checking SKUs first. """
    try:
        output_wb = load_workbook(output_excel_path)
        output_ws = output_wb.active
    except FileNotFoundError:
        output_wb = Workbook()
        output_ws = output_wb.active
        output_ws.append(["Variant SKU", "Description"])
        output_ws.cell(row=1, column=3).value = "Variant SKU"
        output_ws.cell(row=1, column=8).value = "Description"


    sku_row_index = {}
    for row in range(2, output_ws.max_row + 1):
        sku = output_ws.cell(row=row, column=3).value
        if sku:
            sku_row_index[sku] = row

    # Read and update descriptions
    for subdir, dirs, files in os.walk(sub_folder1):
        sku = os.path.basename(subdir)
        description_file = os.path.join(subdir, 'Textprepared.txt')
        if os.path.exists(description_file):
            with open(description_file, 'r', encoding='utf-8') as file:
                description = file.read()
                if sku in sku_row_index:

                    row_index = sku_row_index[sku]
                    output_ws.cell(row=row_index, column=8).value = description
                else:

                    new_row = output_ws.max_row + 1
                    output_ws.cell(row=new_row, column=3).value = sku
                    output_ws.cell(row=new_row, column=8).value = description


    output_wb.save(output_excel_path)
    print("Excel file updated with descriptions.")

def main():
    input_excel_path = r'C:\Project python completed version\(ES) Robinsons 2nd Batch.xlsx'
    output_excel_path = r'C:\Project python completed version\SKU Filled.xlsx'
    sub_folder1 = r'C:\Project python completed version\SKU folder'
    search_url = 'https://www.tefal.com.sg/search'

    sku_list = load_skus_from_excel(input_excel_path)
    found, not_found = check_sku_existence(sku_list, search_url, sub_folder1)
    extract_and_clean_data(sub_folder1)
    update_excel_with_description(output_excel_path, sub_folder1)

if __name__ == "__main__":
    main()